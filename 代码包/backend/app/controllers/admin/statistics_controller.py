"""数据统计 - API328~330"""
import os, uuid, tempfile
from flask import Blueprint, request, send_file
from ...common.response import success, fail
from ...common.auth import admin_required
from ...common.log_helper import log
from ...services.statistics_service import get_overview
from ...models.delivery_record import DeliveryRecord
from ...models.user import User
from ...models.device import Device
from ...extensions import db
from sqlalchemy import func
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

bp = Blueprint("admin_statistics", __name__)

def _comm_filter(query):
    """社区过滤辅助"""
    c = request.args.get("community", "")
    if c:
        return query.filter(DeliveryRecord.deviceId.like(f'{c}%'))
    return query

def _dev_filter(query):
    c = request.args.get("community", "")
    if c:
        return query.filter(Device.deviceId.like(f'{c}%'))
    return query

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center")

def _write_header(ws, headers):
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = HEADER_FONT; c.fill = HEADER_FILL; c.alignment = HEADER_ALIGN

@bp.route("/statistics/overview", methods=["GET"])
@admin_required
def overview():
    community = request.args.get("community", "")
    return success(get_overview(community))

@bp.route("/statistics/delivery", methods=["GET"])
@admin_required
def delivery_stats():
    q = DeliveryRecord.query
    start = request.args.get("startTime", "")
    if start:
        try: q = q.filter(DeliveryRecord.deliveryTime >= datetime.fromisoformat(start))
        except ValueError: pass
    end = request.args.get("endTime", "")
    if end:
        try: q = q.filter(DeliveryRecord.deliveryTime <= datetime.fromisoformat(end))
        except ValueError: pass
    area = request.args.get("area", "")
    if area: q = q.filter(DeliveryRecord.deviceId.contains(area[:3]))
    total = q.count()
    correct = q.filter(DeliveryRecord.isCorrect == True).count()
    points = db.session.query(func.sum(DeliveryRecord.pointChange)).scalar() or 0
    return success({
        "totalDeliveryCount": total, "correctCount": correct,
        "incorrectCount": total - correct,
        "correctRate": round(correct / total, 2) if total > 0 else 0,
        "totalPointsAwarded": int(points) if points > 0 else 0
    })

@bp.route("/statistics/building-compare", methods=["GET"])
@admin_required
def building_compare():
    """各栋分类正确率对比"""
    devices = _dev_filter(Device.query).all()
    bld_map = {}
    for d in devices:
        parts = d.deviceId.split('-')
        if len(parts) >= 2:
            bld = f'{int(parts[1])}栋'
            if bld not in bld_map: bld_map[bld] = []
            bld_map[bld].append(d.deviceId)

    now = datetime.utcnow()
    days30 = now - timedelta(days=30)
    result = []
    for bld in sorted(bld_map.keys(), key=lambda x: int(x.replace('栋',''))):
        dids = bld_map[bld]
        area = 'A区' if int(bld.replace('栋','')) <= 6 else 'B区'
        total = DeliveryRecord.query.filter(DeliveryRecord.deviceId.in_(dids)).count()
        correct = DeliveryRecord.query.filter(DeliveryRecord.deviceId.in_(dids), DeliveryRecord.isCorrect==True).count()
        rate = round(correct/total*100, 1) if total > 0 else 0
        # 参与率 = 近30天活跃用户 / 历史总用户（该楼栋）
        active_users = db.session.query(func.count(func.distinct(DeliveryRecord.userId))).filter(
            DeliveryRecord.deviceId.in_(dids), DeliveryRecord.deliveryTime >= days30).scalar() or 0
        all_users = db.session.query(func.count(func.distinct(DeliveryRecord.userId))).filter(
            DeliveryRecord.deviceId.in_(dids)).scalar() or 1
        participation = round(active_users / all_users * 100, 1)
        pts = db.session.query(func.sum(DeliveryRecord.pointChange)).filter(DeliveryRecord.deviceId.in_(dids)).scalar() or 0
        result.append({'building': bld, 'area': area, 'total': total, 'correct': correct,
                       'rate': rate, 'participation': participation, 'points': int(pts)})
    return success(result)

@bp.route("/statistics/daily-trend", methods=["GET"])
@admin_required
def daily_trend():
    """近 30 天每日投放趋势"""
    from datetime import timedelta
    now = datetime.utcnow()
    trend = []
    for day in range(29, -1, -1):
        d_start = (now - timedelta(days=day)).replace(hour=0, minute=0, second=0, microsecond=0)
        d_end = d_start + timedelta(days=1)
        total = _comm_filter(DeliveryRecord.query.filter(DeliveryRecord.deliveryTime >= d_start, DeliveryRecord.deliveryTime < d_end)).count()
        correct = _comm_filter(DeliveryRecord.query.filter(DeliveryRecord.deliveryTime >= d_start, DeliveryRecord.deliveryTime < d_end, DeliveryRecord.isCorrect==True)).count()
        rate = round(correct/total*100, 1) if total > 0 else 0
        trend.append({'date': d_start.strftime('%m/%d'), 'total': total, 'correct': correct, 'rate': rate})
    return success(trend)

@bp.route("/statistics/category-breakdown", methods=["GET"])
@admin_required
def category_breakdown():
    """四大类垃圾分类正确率与分布"""
    types = [('recyclable','可回收物'), ('kitchen','厨余垃圾'), ('hazardous','有害垃圾'), ('other','其他垃圾')]
    result = []
    for pt, pn in types:
        total = _comm_filter(DeliveryRecord.query.filter_by(parentType=pt)).count()
        correct = _comm_filter(DeliveryRecord.query.filter_by(parentType=pt, isCorrect=True)).count()
        rate = round(correct/total*100, 1) if total > 0 else 0
        result.append({'type': pt, 'name': pn, 'total': total, 'correct': correct, 'rate': rate})
    return success(result)

@bp.route("/statistics/month-compare", methods=["GET"])
@admin_required
def month_compare():
    """本月 vs 上月环比"""
    from datetime import timedelta
    now = datetime.utcnow()
    this_month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    last_month_start = (this_month_start - timedelta(days=1)).replace(day=1)
    def calc(start, end):
        total = _comm_filter(DeliveryRecord.query.filter(DeliveryRecord.deliveryTime >= start, DeliveryRecord.deliveryTime < end)).count()
        correct = _comm_filter(DeliveryRecord.query.filter(DeliveryRecord.deliveryTime >= start, DeliveryRecord.deliveryTime < end, DeliveryRecord.isCorrect == True)).count()
        return total, correct, round(correct/total*100, 1) if total > 0 else 0
    t1, c1, r1 = calc(this_month_start, now)
    t2, c2, r2 = calc(last_month_start, this_month_start)
    return success({'thisMonth': {'total': t1, 'correct': c1, 'rate': r1}, 'lastMonth': {'total': t2, 'correct': c2, 'rate': r2}})

STAT_REPORT_TYPES = {
    'monthly': '月度分类报告',
    'delivery': '投放数据明细',
    'device': '设备状态报表',
    'community': '社区对比报表',
}

@bp.route("/statistics/export", methods=["GET"])
@admin_required
def export_data():
    rtype = request.args.get("type", "delivery")
    wb = Workbook()
    now = datetime.utcnow()
    month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)

    if rtype == 'monthly':
        ws = wb.active; ws.title = "月度汇总"
        _write_header(ws, ["指标","数值","说明"])
        total = DeliveryRecord.query.filter(DeliveryRecord.deliveryTime >= month_start).count()
        correct = DeliveryRecord.query.filter(DeliveryRecord.deliveryTime >= month_start, DeliveryRecord.isCorrect == True).count()
        rate = round(correct/total*100, 1) if total > 0 else 0
        rows_data = [
            ("本月投放总量", total, "次"),
            ("正确投放", correct, "次"),
            ("分类正确率", f"{rate}%", "城管达标线85%"),
            ("在线设备", Device.query.filter_by(onlineStatus="online").count(), "台"),
            ("总设备", Device.query.count(), "台"),
            ("注册用户", User.query.filter_by(userType="resident").count(), "人"),
        ]
        for i, (label, val, note) in enumerate(rows_data, 2):
            ws.cell(row=i, column=1, value=label); ws.cell(row=i, column=2, value=val); ws.cell(row=i, column=3, value=note)

        ws2 = wb.create_sheet("各社区正确率")
        _write_header(ws2, ["社区","总投放","正确","正确率"])
        communities = [('虎溪花园','虎溪'),('学府悦园','学府'),('康居西城','康居'),('龙湖U城','龙湖'),('金科廊桥水乡','金科'),('富力城','富力'),('恒大未来城','恒大'),('融创文旅城','融创')]
        for i, (name, prefix) in enumerate(communities, 2):
            t = DeliveryRecord.query.filter(DeliveryRecord.deviceId.like(f'{prefix}%'), DeliveryRecord.deliveryTime >= month_start).count()
            c = DeliveryRecord.query.filter(DeliveryRecord.deviceId.like(f'{prefix}%'), DeliveryRecord.deliveryTime >= month_start, DeliveryRecord.isCorrect == True).count()
            r = round(c/t*100, 1) if t > 0 else 0
            ws2.cell(row=i, column=1, value=name); ws2.cell(row=i, column=2, value=t)
            ws2.cell(row=i, column=3, value=c); ws2.cell(row=i, column=4, value=f"{r}%")

    elif rtype == 'delivery':
        ws = wb.active; ws.title = "投放记录"
        _write_header(ws, ["记录ID","设备ID","用户ID","垃圾品类","所属大类","是否正确","积分变动","投放时间"])
        for row, r in enumerate(DeliveryRecord.query.order_by(DeliveryRecord.deliveryTime.desc()).limit(5000).all(), 2):
            ws.cell(row=row, column=1, value=r.recordId)
            ws.cell(row=row, column=2, value=r.deviceId); ws.cell(row=row, column=3, value=r.userId)
            ws.cell(row=row, column=4, value=r.garbageCategory); ws.cell(row=row, column=5, value=r.parentType)
            ws.cell(row=row, column=6, value="是" if r.isCorrect else "否")
            ws.cell(row=row, column=7, value=r.pointChange)
            ws.cell(row=row, column=8, value=r.deliveryTime.isoformat() if r.deliveryTime else "")

    elif rtype == 'device':
        ws = wb.active; ws.title = "设备状态"
        _write_header(ws, ["设备ID","设备名称","类型","区域","位置","在线状态","满溢度","最后在线时间"])
        for row, d in enumerate(Device.query.order_by(Device.onlineStatus, Device.lastOnlineTime.desc()).all(), 2):
            ws.cell(row=row, column=1, value=d.deviceId); ws.cell(row=row, column=2, value=d.deviceName)
            ws.cell(row=row, column=3, value=d.boxCategory); ws.cell(row=row, column=4, value=d.area)
            ws.cell(row=row, column=5, value=d.location)
            ws.cell(row=row, column=6, value={'online':'在线','offline':'离线','fault':'故障'}.get(d.onlineStatus, d.onlineStatus))
            ws.cell(row=row, column=7, value=f"{d.fullRate:.0%}" if d.fullRate else "0%")
            ws.cell(row=row, column=8, value=d.lastOnlineTime.isoformat() if d.lastOnlineTime else "")

    else:  # community
        ws = wb.active; ws.title = "社区对比"
        _write_header(ws, ["社区","本月投放","正确","正确率","在线设备","总设备"])
        communities = [('虎溪花园','虎溪'),('学府悦园','学府'),('康居西城','康居'),('龙湖U城','龙湖'),('金科廊桥水乡','金科'),('富力城','富力'),('恒大未来城','恒大'),('融创文旅城','融创')]
        for i, (name, prefix) in enumerate(communities, 2):
            t = DeliveryRecord.query.filter(DeliveryRecord.deviceId.like(f'{prefix}%'), DeliveryRecord.deliveryTime >= month_start).count()
            c = DeliveryRecord.query.filter(DeliveryRecord.deviceId.like(f'{prefix}%'), DeliveryRecord.deliveryTime >= month_start, DeliveryRecord.isCorrect == True).count()
            r = round(c/t*100, 1) if t > 0 else 0
            online = Device.query.filter(Device.deviceId.like(f'{prefix}%'), Device.onlineStatus == 'online').count()
            total_dev = Device.query.filter(Device.deviceId.like(f'{prefix}%')).count()
            ws.cell(row=i, column=1, value=name); ws.cell(row=i, column=2, value=t)
            ws.cell(row=i, column=3, value=c); ws.cell(row=i, column=4, value=f"{r}%")
            ws.cell(row=i, column=5, value=online); ws.cell(row=i, column=6, value=total_dev)

    rpt_name = STAT_REPORT_TYPES.get(rtype, '报表')
    filename = f"{rpt_name}_{uuid.uuid4().hex[:6]}.xlsx"
    filepath = os.path.join(tempfile.gettempdir(), filename)
    wb.save(filepath)
    log('statistics_export', None, f'导出{rpt_name}')
    return success({"exportUrl": f"/api/v1/admin/statistics/download/{filename}", "reportName": rpt_name})

@bp.route("/statistics/download/<filename>", methods=["GET"])
def download_file(filename):
    filepath = os.path.join(tempfile.gettempdir(), filename)
    if not os.path.exists(filepath):
        return fail(404, "文件不存在或已过期")
    return send_file(filepath, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name="report.xlsx")
